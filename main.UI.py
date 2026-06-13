import tkinter as tk
from tkinter import messagebox
from tkinter import simpledialog
from tkinter import ttk
from tkinter import filedialog
from tkinter import font as tkfont
import ctypes
import platform
import json
import os
import time
import random
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from rich.console import Console
from openpyxl import load_workbook
import xlrd
import webbrowser as web
import pyglet
import winreg
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import sys
from typing import Optional, Dict, Any
import requests as req
import shutil
import subprocess
import sys

CONSOLE = Console()
reg_path = r"Software\Microsoft\Windows\CurrentVersion\Themes\Personalize"

def get_system_theme():
    try:
        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, reg_path)
        theme_value = winreg.QueryValueEx(key, "AppsUseLightTheme")[0]
        if theme_value == 1:
            return "Light"
        elif theme_value == 0:
            return "Dark"
        else:
            return "Unknown"
            
    except FileNotFoundError:
        return "Registry key not found"
    except Exception as e:
        return str(e)

def timer():
    def wrapper(func):
        def inner(*args, **kwargs):
            start_time = time.time()
            result = func(*args, **kwargs)
            end_time = time.time()
            CONSOLE.print(f"[*] {func.__name__} executed in {end_time - start_time:.4f} seconds", style="#00bb00")
            return result
        return inner
    return wrapper

def _read_font_family_name(fontfile):
    try:
        with open(fontfile, "rb") as f:
            data = f.read()
        if len(data) < 12:
            return None
        numTables = int.from_bytes(data[4:6], "big")

        offset = 12
        name_offset = None
        for i in range(numTables):
            tag = data[offset:offset+4]
            table_offset = int.from_bytes(data[offset+8:offset+12], "big")
            if tag == b"name":
                name_offset = table_offset
                break
            offset += 16
        if name_offset is None:
            return None
        if name_offset + 6 > len(data):
            return None
        count = int.from_bytes(data[name_offset+2:name_offset+4], "big")
        stringOffset = int.from_bytes(data[name_offset+4:name_offset+6], "big")
        records_offset = name_offset + 6
        candidates = []
        for i in range(count):
            rec_off = records_offset + i*12
            platformID = int.from_bytes(data[rec_off:rec_off+2], "big")
            nameID = int.from_bytes(data[rec_off+6:rec_off+8], "big")
            length = int.from_bytes(data[rec_off+8:rec_off+10], "big")
            offset_in_storage = int.from_bytes(data[rec_off+10:rec_off+12], "big")
            if nameID not in (16, 1):
                continue
            string_pos = name_offset + stringOffset + offset_in_storage
            if string_pos + length > len(data):
                continue
            raw = data[string_pos:string_pos+length]
            try:
                if platformID in (0, 3):
                    text = raw.decode("utf-16-be")
                else:
                    text = raw.decode("latin-1")
            except Exception:
                text = None
            if not text:
                continue
            score = 0
            if nameID == 16:
                score += 20
            if nameID == 1:
                score += 10
            if platformID == 3:
                score += 5
            if platformID == 0:
                score += 4
            candidates.append((score, text))
        if not candidates:
            return None
        candidates.sort(key=lambda x: x[0], reverse=True)
        return candidates[0][1]
    except Exception:
        return None

def fontloader(fontpath):
    if not fontpath:
        return "Microsoft YaHei"
    try:
        if not os.path.exists(fontpath):
            return fontpath
        try:
            pyglet.font.add_file(fontpath)
        except Exception:
            pass
        try:
            if platform.system() == "Windows":
                FR_PRIVATE = 0x10
                AddFontResourceEx = ctypes.windll.gdi32.AddFontResourceExW
                AddFontResourceEx(str(fontpath), FR_PRIVATE, 0)
        except Exception:
            pass
        family = _read_font_family_name(fontpath)
        if family:
            return family
    except Exception as e:
        CONSOLE.print(f"[!] fontloader error: {e}", style="#bb0000")

    return "Microsoft YaHei"

class GitHubReleaseChecker:
    def __init__(self, owner: str, repo: str):
        self.owner = owner
        self.repo = repo
        self.base_url = f"https://api.github.com/repos/{owner}/{repo}"
        self.headers = {
            "Accept": "application/vnd.github.v3+json",
            "User-Agent": "Python-GitHub-Release-Checker"
        }

    def get_latest_release(self) -> Optional[Dict[str, Any]]:
        url = f"{self.base_url}/releases/latest"
        try:
            response = requests.get(url, headers=self.headers, timeout=10)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            CONSOLE.print(f"请求失败: {e}", style="#bb0000")
            return None
        except json.JSONDecodeError:
            CONSOLE.print("解析 JSON 响应失败", style="#bb0000")
            return None

    def get_all_releases(self, per_page: int = 30) -> list:
        url = f"{self.base_url}/releases"
        releases = []
        try:
            response = requests.get(url, headers=self.headers, params={"per_page": per_page}, timeout=10)
            response.raise_for_status()
            releases = response.json()
        except requests.exceptions.RequestException as e:
            CONSOLE.print(f"请求失败: {e}", style="#bb0000")
        return releases

    def parse_version_info(self, release_data: Dict[str, Any]) -> Dict[str, str]:
        if not release_data:
            return {}
        
        return {
            "tag_name": release_data.get("tag_name", "Unknown"),
            "name": release_data.get("name", "Unnamed Release"),
            "published_at": release_data.get("published_at", "Unknown Date"),
            "html_url": release_data.get("html_url", ""),
            "body_preview": release_data.get("body", "")[:200] + "..." if release_data.get("body") else "No description"
        }

class Explain:
    def __init__(self, parent, text, control, font="Microsoft YaHei", size=12, padx=0, pady=4):
        self.parent = parent
        self.text = text
        self.control = control
        self.font = font
        self.size = size
        self.padx = padx
        self.pady = pady

        self.label = tk.Label(
            parent,
            text=text,
            font=(font, size),
            fg="#555"
        )
        self._place_label()

    def _place_label(self):
        self.parent.update_idletasks()
        if not self.control.winfo_ismapped():
            self.parent.after(10, self._place_label)
            return

        x = self.control.winfo_x()
        y = self.control.winfo_y()
        self.label.place(x=x + self.padx, y=y - self.label.winfo_reqheight() - self.pady)

class MainUI():
    def __init__(self):
        self.showInfo()
#------------------Init Variable-----------------------#
        self.latestPos = {"x": 0, "y": 0}
        self.choiseList = []
        self.maxcount = 1
        self.current_file_path = None
        self.countVar = None
        self.roll_job = None
        self.rolling = False
        self.roll_target_count = 1
        self.choice_folder = "assets/choicefile"
        self.font_folder = "assets/font"
        os.makedirs(self.choice_folder, exist_ok=True)
        os.makedirs(self.font_folder, exist_ok=True)
        self.font_file = os.path.join(self.font_folder, "MicrosoftYaHei.ttc")
        self.font = fontloader(self.font_file)
        self.url = "https://github.com/mememe2012/ChoiceUI"
        self.version = "1.3.1.1"

        self.language_file = "English.json"
        self.language_data = {}
        self.theme_key = "light"
        self.themes = json.loads(open("assets/theme/theme.json", "r", encoding="utf-8").read())
#---------------------Init UI--------------------------#
        self.root = tk.Tk()
        self.style = ttk.Style(self.root)
        if "clam" in self.style.theme_names():
            try:
                self.style.theme_use("clam")
            except Exception:
                pass
        self.countVar = tk.IntVar(self.root, value=1)
        self.root.resizable(False, False)

        self.LoadSetting()
        self.Controls()
        self.ApplyLanguage()
        self.ApplyTheme()

        if hasattr(self, 'pending_load_file') and self.pending_load_file:
            self.LoadFile(self.pending_load_file)

        self.root.iconbitmap("assets/icon.ico")
#---------------------File----------------------------#
        if os.path.exists("setuptmp"):
            CONSOLE.print("Cleaning up temporary files...", style="#00bb00")
            shutil.rmtree("setuptmp", ignore_errors=True)
        try:os.mkdir("setuptmp")
        except Exception:pass
        if os.path.exists("setting.json"):
            shutil.copy("setting.json", "setuptmp/setting.json")
            os.remove("setting.json")
#------------------Check UPDATE-----------------------#
        gcu = threading.Thread(target=lambda: self.checkupdate(self.version))
        gcu.daemon = True
        gcu.start()
#=====================================================#
        self.threadPrograss(self.UIposChanged)

        self.root.mainloop()

    def cancel_download(self):
        if messagebox.askyesno(self.tr("取消下载"), self.tr("是否取消当前下载？")):
            self.download_cancelled = True
            if self.rolling:
                self.rolling = False
            if self.roll_job:
                self.root.after_cancel(self.roll_job)
                self.roll_job = None
            if hasattr(self, 'progresswindow') and self.progresswindow.winfo_exists():
                self.progresswindow.destroy()

    def downloadFile(self, url, save_path):
        self.progresswindow = tk.Toplevel(self.root)
        self.progresswindow.title(self.tr("下载"))
        self.progresswindow.resizable(False, False)
        self.progresswindow.geometry("400x100")
        self.progresswindow.iconbitmap("assets/icon.ico")
        self.progresswindow.config(bg=self.themes[self.theme_key]["bg"])

        self.progressbar = ttk.Progressbar(self.progresswindow, style="Green.Horizontal.TProgressbar", mode="determinate", maximum=100, value=0)
        self.progressbar.place(x=20, y=20, width=360, height=30)

        self.progressLabel = tk.Label(self.progresswindow, text=self.tr("Requesting..."), font=(self.font, 10), bg=self.themes[self.theme_key]["bg"], fg=self.themes[self.theme_key]["fg"])
        self.progressLabel.place(x=20, y=60)

        self.speedLabel = tk.Label(self.progresswindow, text="0.00 MiB/s", font=(self.font, 10), bg=self.themes[self.theme_key]["bg"], fg=self.themes[self.theme_key]["fg"])
        self.speedLabel.place(x=20, y=80)

        self.progresswindow.grab_set()

        self.progresswindow.protocol("WM_DELETE_WINDOW", self.cancel_download)

        self.download_cancelled = False
        start_time = time.time()

        def update_progress_ui(downloaded_bytes: int, total_bytes: Optional[int]):
            elapsed = max(time.time() - start_time, 0.001)
            speed = downloaded_bytes / elapsed
            speed_text = f"{speed / 1024 / 1024:.2f} MiB/s"
            self.speedLabel.config(text="{speed}".format(speed=speed_text))
            if total_bytes:
                percent = min(downloaded_bytes / total_bytes * 100, 100)
                self.progressLabel.config(text=f"{downloaded_bytes / 1024 / 1024:.2f} MiB/{total_bytes / 1024 / 1024:.2f} MiB {percent:.2f}%")
                self.progressbar['value'] = percent
            else:
                self.progressLabel.config(text=f"{downloaded_bytes / 1024 / 1024:.2f} MiB")
                self.progressbar['value'] = 0

        def download_segment(index: int, start: int, end: int, total_bytes: int, headers: Dict[str, str], adapter: HTTPAdapter, temp_path: str, downloaded_lock: threading.Lock,  downloaded_state: Dict[str, int]):
            part_path = f"{temp_path}.part{index}"
            local_headers = headers.copy()
            local_headers["Range"] = f"bytes={start}-{end}"
            with req.Session() as segment_session:
                segment_session.mount("https://", adapter)
                segment_session.mount("http://", adapter)
                with segment_session.get(url, headers=local_headers, stream=True, timeout=(10, 60), allow_redirects=True) as segment_response:
                    segment_response.raise_for_status()
                    with open(part_path, "wb") as part_file:
                        for chunk in segment_response.iter_content(chunk_size=8192):
                            if self.download_cancelled:
                                raise RuntimeError("Download cancelled")
                            if chunk:
                                part_file.write(chunk)
                                with downloaded_lock:
                                    downloaded_state["bytes"] += len(chunk)
                                    current = downloaded_state["bytes"]
                                self.root.after(0, update_progress_ui, current, total_bytes)
            return part_path

        try:
            session = req.Session()
            retries = Retry(
                total=5,
                backoff_factor=0.6,
                status_forcelist=(429, 500, 502, 503, 504),
                allowed_methods=("HEAD", "GET", "OPTIONS")
            )
            adapter = HTTPAdapter(max_retries=retries)
            session.mount("https://", adapter)
            session.mount("http://", adapter)

            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
                "Accept": "application/octet-stream"
            }

            CONSOLE.print(f"正在下载文件: {url}", style="#00bb00")
            total_length = None
            accept_ranges = False
            temp_path = save_path + ".part"

            try:
                with session.head(url, headers=headers, allow_redirects=True, timeout=(10, 60)) as head_response:
                    head_response.raise_for_status()
                    total_length = head_response.headers.get("content-length")
                    accept_ranges = head_response.headers.get("accept-ranges", "").lower() == "bytes"
            except requests.exceptions.RequestException:
                pass

            if total_length is not None:
                total_length = int(total_length)
                self.progressLabel.config(text=f"File Size: {total_length / 1024:.2f} KB")
            else:
                self.progressLabel.config(text=self.tr("File Size: Unknown"))

            if total_length and accept_ranges:
                self.progressbar.config(mode="determinate", maximum=100, value=0)
                max_workers = min(4, (os.cpu_count() or 2) or 1)
                segment_count = max_workers
                segment_size = total_length // segment_count
                downloaded_lock = threading.Lock()
                downloaded_state = {"bytes": 0}
                part_paths = []
                ranges = []
                for index in range(segment_count):
                    start = index * segment_size
                    end = total_length - 1 if index == segment_count - 1 else ((index + 1) * segment_size - 1)
                    ranges.append((index, start, end))

                with ThreadPoolExecutor(max_workers=segment_count) as executor:
                    future_to_part = {
                        executor.submit(download_segment, idx, start, end, total_length, headers, adapter, temp_path, downloaded_lock, downloaded_state): idx
                        for idx, start, end in ranges
                    }
                    for future in as_completed(future_to_part):
                        if self.download_cancelled:
                            raise RuntimeError("Download cancelled")
                        part_paths.append(future.result())

                with open(temp_path, "wb") as output_file:
                    for index in range(len(part_paths)):
                        part_path = f"{temp_path}.part{index}"
                        with open(part_path, "rb") as part_file:
                            output_file.write(part_file.read())
                        try:
                            os.remove(part_path)
                        except Exception:
                            pass

            else:
                with session.get(url, headers=headers, stream=True, timeout=(10, 60), allow_redirects=True) as response:
                    response.raise_for_status()
                    total_length = response.headers.get("content-length")
                    if total_length is not None:
                        total_length = int(total_length)
                        self.progressbar.config(mode="determinate", maximum=100, value=0)
                    else:
                        self.progressbar.config(mode="indeterminate")
                        self.progressbar.start(10)

                    downloaded = 0
                    with open(temp_path, "wb") as f:
                        for chunk in response.iter_content(chunk_size=8192):
                            if self.download_cancelled:
                                raise RuntimeError("Download cancelled")
                            if chunk:
                                f.write(chunk)
                                downloaded += len(chunk)
                                if total_length:
                                    percent = min(downloaded / total_length * 100, 100)
                                    CONSOLE.print(f"下载进度: {percent:.2f}%", style="#00bb00", end="\r")
                                    self.root.after(0, update_progress_ui, downloaded, total_length)
                                else:
                                    self.root.after(0, update_progress_ui, downloaded, None)
                    if total_length:
                        self.progressbar['value'] = 100

                self.progressbar.destroy()
                self.speedLabel.destroy()

                self.textlabel = tk.Label(self.progresswindow, text="Loading...", font=(self.font, 10), bg=self.themes[self.theme_key]["bg"], fg=self.themes[self.theme_key]["fg"])
                self.textlabel.place(x=20, y=60)

            if self.download_cancelled:
                raise RuntimeError("Download cancelled")
            
            self.progressLabel.config(text="Saveing...")

            os.replace(temp_path, save_path)
            CONSOLE.print(f"下载完成，已保存到: {save_path}", style="#00bb00")
            
            self.loadupdatefile()

        except RuntimeError as e:
            if str(e) == "Download cancelled":
                CONSOLE.print("下载已取消。", style="#bb0000")
            else:
                CONSOLE.print(f"Error: {e}", style="#bb0000")
                messagebox.showerror(self.tr("错误"), f"{self.tr('下载文件失败')}: {e}")
        except requests.exceptions.RequestException as e:
            CONSOLE.print(f"下载请求失败: {e}", style="#bb0000")
            messagebox.showerror(self.tr("错误"), f"{self.tr('下载文件失败')}: {e}")
        except Exception as e:
            CONSOLE.print(f"Error: {e}", style="#bb0000")
            messagebox.showerror(self.tr("错误"), f"{self.tr('下载文件失败')}: {e}")
        finally:
            if hasattr(self, 'progresswindow') and self.progresswindow.winfo_exists():
                self.progresswindow.destroy()

    def loadupdatefile(self):
        try:
            if os.path.exists("assets/setting.json"):
                shutil.copy("assets/setting.json", "setting.json")
            CONSOLE.print("已准备好安装程序。正在启动...", style="#00bb00")
        except Exception as e:
            CONSOLE.print(f"Failed to copy setting.json: {e}", style="#bb0000")
        process = subprocess.Popen(
            ["setuptmp/setup.exe"], 
            stdout=subprocess.DEVNULL, 
            stderr=subprocess.DEVNULL,
            creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == 'win32' else 0
        )
        CONSOLE.print("正在启动安装程序...", style="#00bb00")
        sys.exit(0)

    def checkupdate(self, version):
        if os.path.exists("setuptmp"):
            CONSOLE.print("正在清理临时文件...", style="#00bb00")
            shutil.rmtree("setuptmp", ignore_errors=True)
        version = "V" + version
        if len(sys.argv) < 3:
            owner = "mememe2012"
            repo = "ChoiceUI"
        else:
            owner = sys.argv[1]
            repo = sys.argv[2]

        CONSOLE.print(f"正在检查 GitHub 仓库: {owner}/{repo} 的最新版本...", style="#00bb00")
        
        checker = GitHubReleaseChecker(owner, repo)
        
        latest_release = checker.get_latest_release()
        
        if latest_release:
            info = checker.parse_version_info(latest_release)
            CONSOLE.print("-" * 40, style="#00bb00")
            CONSOLE.print(f"最新标签 (Tag): {info['tag_name']}", style="#00bb00")
            CONSOLE.print(f"发布名称:       {info['name']}", style="#00bb00")
            CONSOLE.print(f"发布时间:       {info['published_at']}", style="#00bb00")
            CONSOLE.print(f"链接:           {info['html_url']}", style="#00bb00")
            CONSOLE.print(f"简介预览:       \n{"="*10+"简介"+"="*10}\n{info['body_preview']}", style="#00bb00")
            CONSOLE.print("-" * 40, style="#00bb00")
            v1 = "".join(info['tag_name'].lstrip('V').split('.'))
            v2 = "".join(version.lstrip('V').split('.'))
            if int(v1) > int(v2):
                CONSOLE.print(f"检测到新版本！当前版本: {version}，最新版本: {info['tag_name']}", style="#00bb00")
                if messagebox.askyesno("更新提示", self.tr("检测到新版本！当前版本: {version}\n最新版本: {info}\n下载链接: {url}\n是否升级？").format(version=version, info=info['tag_name'], url=info['html_url'])):
                    dowloadurl = f"{"/".join(info['html_url'].split('/')[:-3])}/releases/download/{info['tag_name']}/ChoiceUI{info['tag_name'].replace('V', '')}-SETUP.exe"
                    try:os.mkdir("setuptmp")
                    except Exception:pass
                    self.downloadFile(dowloadurl, "setuptmp/setup.exe")
            else:
                CONSOLE.print(f"当前版本 {version} 已是最新版本。", style="#00bb00")
                messagebox.showinfo(self.tr("检查更新"), self.tr("当前版本 {version} 已是最新版本。").format(version=version))
        else:
            CONSOLE.print("未能获取最新版本信息。请检查仓库名称是否正确或网络连接。", style="#bb0000")
            messagebox.showerror("错误",self.tr("未能获取最新版本信息。请检查仓库名称是否正确或网络连接。"))

    def Controls(self):
        self.showLabel = tk.Label(self.root, text=self.tr("等待抽取"), font=(self.font, 20))
        self.showLabel.place(x=400, y=10, anchor="n")

        self.startButton = ttk.Button(self.root, text=self.tr("开始抽奖"), command=self.Choice)
        self.startButton.place(x=300, y=60, anchor="n")

        self.setButton = ttk.Button(self.root, text=self.tr("设置选项"), command=self.Setting)
        self.setButton.place(x=500, y=60, anchor="n")

        self.countSpinBox = ttk.Spinbox(self.root, from_=1, to=100, textvariable=self.countVar, width=5)
        self.countSpinBox.place(x=300, y=150, anchor="n", width=100)
        Explain(self.root, self.tr("抽取个数"), self.countSpinBox)

        self.countInfoLabel = tk.Label(self.root, text=self.tr("名单数量：0"), font=(self.font, 10), fg="#333")
        self.countInfoLabel.place(x=420, y=150, anchor="n")

        self.loadedFileLabel = tk.Label(self.root, text=self.tr("当前名单：无"), font=(self.font, 10), fg="#333")
        self.loadedFileLabel.place(x=420, y=180, anchor="n")

        self.MainMenu = tk.Menu(self.root)
        self.root.config(menu=self.MainMenu)

        self.codeMenu = tk.Menu(self.MainMenu, tearoff=0)
        self.codeMenu.add_command(label="源代码", command=self.codeSource)
        self.codeMenu.add_command(label="关于", command=self.about)
        self.MainMenu.add_cascade(label="软件信息", menu=self.codeMenu)

        self.SettingMenu = tk.Menu(self.MainMenu, tearoff=0)
        self.SettingMenu.add_command(label="主题", command=self.theme)
        self.SettingMenu.add_command(label="语言", command=self.lang)
        self.SettingMenu.add_command(label="字体", command=self.font_setting)
        self.SettingMenu.add_command(label="检查更新", command=lambda: threading.Thread(target=self.checkupdate, args=(self.version,), daemon=True).start())
        self.MainMenu.add_cascade(label="设置", menu=self.SettingMenu)

        self.UpdateCountLimit()

    def theme(self):
        self.themeroot = tk.Toplevel(self.root)
        self.themeroot.title(self.tr("主题"))
        self.themeroot.resizable(False, False)
        self.themeroot.geometry("300x300")
        self.themeroot.iconbitmap("assets/icon.ico")
        self.themeroot.config(bg=self.themes[self.theme_key]["bg"])

        theme_listbox = tk.Listbox(self.themeroot, width=24, height=6, activestyle="dotbox")
        theme_listbox.place(x=20, y=20)
        for index, (key, data) in enumerate(self.themes.items()):
            theme_listbox.insert(tk.END, self.tr(data["name"]))
            if key == self.theme_key:
                theme_listbox.selection_set(index)

        def apply_selected_theme():
            selection = theme_listbox.curselection()
            if not selection:
                return
            index = selection[0]
            key = list(self.themes.keys())[index]
            self.theme_key = key
            self.ApplyTheme()
            self.SaveSetting("Theme", key)
            messagebox.showinfo(self.tr("主题"), self.tr("主题已应用。"))
            self.themeroot.destroy()

        apply_button = ttk.Button(self.themeroot, text=self.tr("应用"), command=apply_selected_theme)
        apply_button.place(x=20, y=170, width=100)
        close_button = ttk.Button(self.themeroot, text=self.tr("取消"), command=self.themeroot.destroy)
        close_button.place(x=140, y=170, width=100)
        tk.Label(self.themeroot, text=self.tr("系统主题:") + get_system_theme(), font=(self.font, 10), bg=self.themes[self.theme_key]["bg"], fg=self.themes[self.theme_key]["fg"]).place(x=20, y=260)
        self.apply_theme_to_window(self.themeroot)

    def lang(self):
        self.langroot = tk.Toplevel(self.root)
        self.langroot.title(self.tr("语言"))
        self.langroot.resizable(False, False)
        self.langroot.geometry("400x300")
        self.langroot.iconbitmap("assets/icon.ico")
        self.langroot.config(bg=self.themes[self.theme_key]["bg"])

        language_listbox = tk.Listbox(self.langroot, width=32, height=8, activestyle="dotbox")
        language_listbox.place(x=20, y=20)
        language_files = self.LoadLanguageFiles()
        for index, (filename, display) in enumerate(language_files):
            language_listbox.insert(tk.END, display)
            if filename == self.language_file:
                language_listbox.selection_set(index)

        def apply_language():
            selection = language_listbox.curselection()
            if not selection:
                return
            filename = language_files[selection[0]][0]
            self.ChangeLanguage(filename)
            messagebox.showinfo(self.tr("语言"), self.tr("语言已切换。"))
            self.langroot.destroy()

        apply_button = ttk.Button(self.langroot, text=self.tr("应用"), command=apply_language)
        apply_button.place(x=20, y=250, width=100)
        close_button = ttk.Button(self.langroot, text=self.tr("取消"), command=self.langroot.destroy)
        close_button.place(x=140, y=250, width=100)
        self.apply_theme_to_window(self.langroot)

    def font_setting(self):
        self.fontroot = tk.Toplevel(self.root)
        self.fontroot.title(self.tr("字体"))
        self.fontroot.resizable(False, False)
        self.fontroot.geometry("650x360")
        self.fontroot.iconbitmap("assets/icon.ico")
        self.fontroot.config(bg=self.themes[self.theme_key]["bg"])

        self.fontLabel = tk.Label(self.fontroot, text=self.tr("当前字体：") + self.GetFontDisplayName(), font=(self.font, 10))
        self.fontLabel.place(x=20, y=20)

        self.fontListbox = tk.Listbox(self.fontroot, height=10, width=40)
        self.fontListbox.place(x=20, y=60)
        self.fontListbox.bind("<<ListboxSelect>>", self.OnFontSelectChange)

        self.previewTitle = tk.Label(self.fontroot, text=self.tr("字体预览"), font=(self.font, 10))
        self.previewTitle.place(x=360, y=20)

        self.previewLabel = tk.Label(self.fontroot, text=self.tr("文本\nAaBbCc\n123"), font=(self.font, 20), bg=self.themes[self.theme_key]["entry_bg"], fg=self.themes[self.theme_key]["fg"], width=20, height=6, anchor="center", justify="center", wraplength=140)
        self.previewLabel.place(x=400, y=60, width=160, height=160)

        self.importFontButton = ttk.Button(self.fontroot, text=self.tr("导入字体"), command=self.ImportFontFile)
        self.importFontButton.place(x=20, y=300, width=120)

        self.selectFontButton = ttk.Button(self.fontroot, text=self.tr("选择字体"), command=self.SelectFontFile)
        self.selectFontButton.place(x=160, y=300, width=120)

        self.deleteFontButton = ttk.Button(self.fontroot, text=self.tr("删除字体"), command=self.DeleteFontFile)
        self.deleteFontButton.place(x=300, y=300, width=120)

        self.RefreshFontList()
        self.UpdateFontPreview(self.font_file)
        self.apply_theme_to_window(self.fontroot)
        self.fontroot.grab_set()

    def RefreshFontList(self):
        self.fontListbox.delete(0, tk.END)
        fonts = self.GetAvailableFonts()
        for index, font_name in enumerate(fonts):
            self.fontListbox.insert(tk.END, font_name)
            font_path = os.path.join(self.font_folder, font_name)
            if self.font_file and os.path.abspath(font_path) == os.path.abspath(self.font_file):
                self.fontListbox.selection_set(index)

    def GetAvailableFonts(self):
        fonts = []
        try:
            for filename in os.listdir(self.font_folder):
                if filename.lower().endswith((".ttf", ".ttc", ".otf")):
                    fonts.append(filename)
        except Exception:
            pass
        return sorted(fonts)

    def SelectFontFile(self):
        selection = self.fontListbox.curselection()
        if not selection:
            messagebox.showwarning(self.tr("提示"), self.tr("请先选择一个字体文件。"))
            return
        filename = self.fontListbox.get(selection[0])
        font_path = os.path.join(self.font_folder, filename)
        self.font_file = font_path
        self.font = self.LoadFont(font_path)
        self.fontLabel.config(text=self.tr("当前字体：") + self.GetFontDisplayName(), font=(self.font, 10))
        self.ApplyFont()
        self.UpdateFontPreview(font_path)
        self.SaveSetting("Font", self.font_file)
        messagebox.showinfo(self.tr("字体"), self.tr("字体已切换。"))
        self.RefreshFontList()

    def ImportFontFile(self):
        path = filedialog.askopenfilename(
            title=self.tr("导入字体"),
            filetypes=[("Font files", "*.ttf *.ttc *.otf"), ("All files", "*")]
        )
        if not path:
            return
        filename = os.path.basename(path)
        target_path = os.path.join(self.font_folder, filename)
        count = 1
        base, ext = os.path.splitext(filename)
        while os.path.exists(target_path):
            target_path = os.path.join(self.font_folder, f"{base}_{count}{ext}")
            count += 1
        try:
            with open(path, "rb") as src, open(target_path, "wb") as dst:
                dst.write(src.read())
            self.RefreshFontList()
            messagebox.showinfo(self.tr("导入字体"), self.tr("字体已导入：{filename}").format(filename=os.path.basename(target_path)))
        except Exception as e:
            CONSOLE.print(f"[!] Failed to import font: {e}", style="#bb0000")
            messagebox.showerror(self.tr("错误"), f"{self.tr('导入字体失败')}: {e}")

    def DeleteFontFile(self):
        selection = self.fontListbox.curselection()
        if not selection:
            messagebox.showwarning(self.tr("提示"), self.tr("请先选择一个字体文件。"))
            return
        filename = self.fontListbox.get(selection[0])
        font_path = os.path.join(self.font_folder, filename)
        if messagebox.askyesno(self.tr("确认删除"), self.tr("是否删除字体文件：{filename}?" ).format(filename=filename)):
            try:
                if self.font_file and os.path.abspath(font_path) == os.path.abspath(self.font_file):
                    self.UnregisterFontFile(self.font_file)
                    self.font_file = None
                    self.font = self.LoadFont(None)
                    self.SaveSetting("Font", self.font_file)
                else:
                    self.UnregisterFontFile(font_path)
                os.remove(font_path)
                self.RefreshFontList()
                self.fontLabel.config(text=self.tr("当前字体：") + self.GetFontDisplayName(), font=(self.font, 10))
                self.ApplyFont()
                self.UpdateFontPreview(self.font_file)
                messagebox.showinfo(self.tr("删除成功"), self.tr("已删除：{filename}").format(filename=filename))
            except Exception as e:
                CONSOLE.print(f"[!] Failed to delete font: {e}", style="#bb0000")
                messagebox.showerror(self.tr("错误"), f"{self.tr('删除字体失败')}: {e}")

    def LoadFont(self, font_value):
        if not font_value:
            CONSOLE.print("[-] No font specified, using default.", style="#bbbb00")
            return fontloader("Microsoft YaHei")
        if os.path.exists(font_value):
            CONSOLE.print(f"[*] Loaded font: {font_value}", style="#00bb00")
            return fontloader(font_value)
        return fontloader("Microsoft YaHei")

    def codeSource(self):
        choice = messagebox.askyesno(self.tr("源代码"), self.tr("是否打开 Choice UI 的 GitHub 仓库？"))
        if choice:
            web.open(self.url)

    def about(self):
        messagebox.showinfo(self.tr("关于"), self.tr("关于信息 v{version}\n作者：mememe2012\nGithub：{url}\nLICENSE: MIT License\n{license}").format(
            version=self.version,
            url=self.url,
            license=open("LICENSE", "r", encoding="utf-8").read()
        ))

    def showInfo(self):
        def show():
            with open("assets/CONTENT.txt", "r", encoding="utf-8") as f:
                info = f.read()
            CONSOLE.print(info, style="#ffff00")

        show()

    def Setting(self):
        class SettingUI():
            def __init__(self, parent):
                self.parent = parent
                self.setroot = tk.Toplevel(parent.root)
                self.setroot.title(parent.tr("设置选项"))
                self.setroot.resizable(False, False)
                self.setroot.iconbitmap("assets/icon.ico")
                self.setroot.geometry("520x360")
                self.setroot.config(bg=parent.themes[parent.theme_key]["bg"])
                self.Setcontrols()
                self.RefreshList()
                parent.apply_theme_to_window(self.setroot)
                self.setroot.grab_set()

            def Setcontrols(self):
                self.importButton = ttk.Button(self.setroot, text=self.parent.tr("导入TXT名单"), command=self.parent.askFile)
                self.importButton.place(x=80, y=20, anchor="n", width=120)

                self.newButton = ttk.Button(self.setroot, text=self.parent.tr("新建空名单"), command=self.CreateEmptyList)
                self.newButton.place(x=200, y=20, anchor="n", width=120)

                self.refreshButton = ttk.Button(self.setroot, text=self.parent.tr("刷新列表"), command=self.RefreshList)
                self.refreshButton.place(x=320, y=20, anchor="n", width=120)

                self.renameButton = ttk.Button(self.setroot, text=self.parent.tr("重命名名单"), command=self.RenameSelectedFile)
                self.renameButton.place(x=440, y=20, anchor="n", width=120)

                self.fileListbox = tk.Listbox(self.setroot, height=10, width=45)
                self.fileListbox.place(x=260, y=70, anchor="n")

                self.loadButton = ttk.Button(self.setroot, text=self.parent.tr("加载名单"), command=self.LoadSelectedFile)
                self.loadButton.place(x=80, y=300, anchor="n", width=120)

                self.editButton = ttk.Button(self.setroot, text=self.parent.tr("编辑名单"), command=self.OpenEditor)
                self.editButton.place(x=200, y=300, anchor="n", width=120)

                self.exportButton = ttk.Button(self.setroot, text=self.parent.tr("导出TXT"), command=self.ExportSelectedFile)
                self.exportButton.place(x=320, y=300, anchor="n", width=120)

                self.deleteButton = ttk.Button(self.setroot, text=self.parent.tr("删除名单"), command=self.DeleteSelectedFile)
                self.deleteButton.place(x=440, y=300, anchor="n", width=120)

            def RefreshList(self):
                self.fileListbox.delete(0, tk.END)
                files = self.parent.LoadChoiceFiles()
                for f in files:
                    self.fileListbox.insert(tk.END, f)

            def GetSelectedFile(self):
                selection = self.fileListbox.curselection()
                if not selection:
                    CONSOLE.print("[-] No file selected.", style="#bbbb00")
                    messagebox.showwarning(self.parent.tr("提示"), self.parent.tr("请先选择一个名单文件。"))
                    return None
                return self.fileListbox.get(selection[0])

            def LoadSelectedFile(self):
                filename = self.GetSelectedFile()
                if not filename:
                    return
                file_path = os.path.join(self.parent.choice_folder, filename)
                self.parent.LoadFile(file_path)
                CONSOLE.print(f"[*] Loaded file: {filename}", style="#00bb00")
                messagebox.showinfo(self.parent.tr("加载成功"), self.parent.tr("已加载名单：{filename}").format(filename=filename))

            def OpenEditor(self):
                filename = self.GetSelectedFile()
                if not filename:
                    return
                file_path = os.path.join(self.parent.choice_folder, filename)
                FileEditor(self.parent, file_path)

            def ExportSelectedFile(self):
                filename = self.GetSelectedFile()
                if not filename:
                    return
                file_path = os.path.join(self.parent.choice_folder, filename)
                save_path = filedialog.asksaveasfilename(
                    title=self.parent.tr("导出TXT"),
                    defaultextension=".txt",
                    filetypes=[("Text files", "*.txt")],
                    initialfile=os.path.splitext(filename)[0] + ".txt"
                )
                if not save_path:
                    return
                try:
                    with open(file_path, "r", encoding="utf-8") as f:
                        names = json.load(f)
                    with open(save_path, "w", encoding="utf-8") as f:
                        for name in names:
                            f.write(str(name).strip() + "\n")
                    CONSOLE.print(f"[*] Exported {len(names)} names to TXT: {save_path}", style="#00bb00")
                    messagebox.showinfo(self.parent.tr("导出成功"), self.parent.tr("已导出到：{path}").format(path=save_path))
                except Exception as e:
                    CONSOLE.print(f"[!] Failed to export TXT: {e}", style="#bb0000")
                    messagebox.showerror(self.parent.tr("错误"), f"{self.parent.tr('导出TXT失败')}: {e}")

            def DeleteSelectedFile(self):
                filename = self.GetSelectedFile()
                if not filename:
                    return
                file_path = os.path.join(self.parent.choice_folder, filename)
                if messagebox.askyesno(self.parent.tr("确认删除"), self.parent.tr("是否删除名单文件：{filename}?").format(filename=filename)):
                    try:
                        os.remove(file_path)
                        self.RefreshList()
                        messagebox.showinfo(self.parent.tr("删除成功"), self.parent.tr("已删除：{filename}").format(filename=filename))
                    except Exception as e:
                        CONSOLE.print(f"[!] Failed to delete file: {e}", style="#bb0000")
                        messagebox.showerror(self.parent.tr("错误"), f"{self.parent.tr('删除文件失败')}: {e}")

            def CreateEmptyList(self):
                new_name = simpledialog.askstring(self.parent.tr("新建空名单"), self.parent.tr("请输入名单文件名（不含扩展名）："), parent=self.setroot)
                if not new_name:
                    return
                new_name = new_name.strip()
                if not new_name:
                    messagebox.showwarning(self.parent.tr("提示"), self.parent.tr("文件名不能为空。"))
                    return
                if not new_name.lower().endswith(".json"):
                    new_name = f"{new_name}.json"
                target_path = os.path.join(self.parent.choice_folder, new_name)
                count = 1
                while os.path.exists(target_path):
                    base, ext = os.path.splitext(new_name)
                    target_path = os.path.join(self.parent.choice_folder, f"{base}_{count}{ext}")
                    count += 1
                try:
                    with open(target_path, "w", encoding="utf-8") as f:
                        json.dump([], f, ensure_ascii=False, indent=2)
                    self.RefreshList()
                    CONSOLE.print(f"[*] Created empty list file: {target_path}", style="#00bb00")
                    messagebox.showinfo(self.parent.tr("新建成功"), self.parent.tr("已创建空名单：{filename}").format(filename=os.path.basename(target_path)))
                except Exception as e:
                    CONSOLE.print(f"[!] Failed to create file: {e}", style="#bb0000")
                    messagebox.showerror(self.parent.tr("错误"), f"{self.parent.tr('创建名单文件失败')}: {e}")

            def RenameSelectedFile(self):
                filename = self.GetSelectedFile()
                if not filename:
                    return
                old_path = os.path.join(self.parent.choice_folder, filename)
                new_name = simpledialog.askstring(self.parent.tr("重命名名单"), self.parent.tr("请输入新的文件名（不含扩展名）："), parent=self.setroot)
                if not new_name:
                    return
                new_name = new_name.strip()
                if not new_name:
                    messagebox.showwarning(self.parent.tr("提示"), self.parent.tr("文件名不能为空。"))
                    return
                if not new_name.lower().endswith('.json'):
                    new_name = f"{new_name}.json"
                target_path = os.path.join(self.parent.choice_folder, new_name)
                if os.path.abspath(target_path) == os.path.abspath(old_path):
                    messagebox.showinfo(self.parent.tr("提示"), self.parent.tr("新旧文件名相同，无需重命名。"))
                    return
                count = 1
                base, ext = os.path.splitext(new_name)
                while os.path.exists(target_path):
                    target_path = os.path.join(self.parent.choice_folder, f"{base}_{count}{ext}")
                    count += 1
                try:
                    os.rename(old_path, target_path)
                    if hasattr(self.parent, 'current_file_path') and os.path.abspath(old_path) == os.path.abspath(self.parent.current_file_path):
                        self.parent.current_file_path = os.path.abspath(target_path)
                        self.parent.UpdateLoadedFileLabel()
                        self.parent.SaveSetting('LoadFile', self.parent.current_file_path)
                    self.RefreshList()
                    CONSOLE.print(f"[*] File renamed: {filename} -> {os.path.basename(target_path)}", style="#00bb00")
                    messagebox.showinfo(self.parent.tr("重命名成功"), self.parent.tr("已将 {old} 重命名为 {new}").format(old=filename, new=os.path.basename(target_path)))
                except Exception as e:
                    CONSOLE.print(f"[!] Failed to rename file: {e}", style="#bb0000")
                    messagebox.showerror(self.parent.tr("错误"), f"{self.parent.tr('重命名失败')}: {e}")

        SettingUI(self)

    def askFile(self):
        path = filedialog.askopenfilename(
            title=self.tr("选择名单文件"),
            filetypes=[("Text files", "*.txt"), ("JSON files", "*.json"), ("All files", "*")]
        )
        if not path:
            return
        self.LoadFile(path)

    def Choice(self):
        if not self.choiseList:
            self.showLabel.config(text=self.tr("没有可抽取的选项"), fg="red")
            return
        if self.rolling:
            return
        count = self.countVar.get()
        if count < 1:
            messagebox.showwarning(self.tr("提示"), self.tr("抽取个数至少为1"))
            return
        if count > len(self.choiseList):
            messagebox.showwarning(self.tr("提示"), self.tr("抽取个数不能超过名单总数"))
            return
        self.roll_target_count = count
        self.rolling = True
        self.startButton.state(["disabled"])
        self.showLabel.config(text=self.tr("正在抽奖..."), fg=self.themes[self.theme_key]["accent"])
        self.roll_start_time = time.time()
        self._roll_animation()

    def _roll_animation(self):
        if not self.rolling:
            return
        elapsed = time.time() - self.roll_start_time
        self.showLabel.config(text=random.choice(self.choiseList), fg=self.themes[self.theme_key]["fg"])
        if elapsed < 3:
            self.roll_job = self.root.after(80, self._roll_animation)
            return
        self.rolling = False
        self.startButton.state(["!disabled"])
        winners = random.sample(self.choiseList, self.roll_target_count)
        result_text = " ".join(winners)
        self.showLabel.config(text=self.tr("中奖：{result}").format(result=result_text), fg=self.themes[self.theme_key]["accent"])

    def threadPrograss(self, *funcs):
        def run():
            for func in funcs:
                func()
        progress = threading.Thread(target=run)
        progress.daemon = True
        progress.start()

    @timer()
    def LoadSetting(self):
        try:
            default_setting = {
                "ScreenPos": {"x": 0, "y": 0},
                "LoadFile": None,
                "Theme": self.theme_key,
                "Language": self.language_file,
                "Font": self.font_file
            }
            if not os.path.exists("assets/setting.json"):
                with open("assets/setting.json", "w", encoding="utf-8") as f:
                    json.dump(default_setting, f, ensure_ascii=False, indent=2)
                setting = default_setting
            else:
                with open("assets/setting.json", "r", encoding="utf-8") as f:
                    setting = json.load(f)
                if not isinstance(setting, dict):
                    CONSOLE.print("[!] Invalid settings format, resetting to defaults.", style="#bbbb00")
                    setting = default_setting
                    with open("assets/setting.json", "w", encoding="utf-8") as f:
                        json.dump(default_setting, f, ensure_ascii=False, indent=2)

            self.latestPos = setting.get("ScreenPos", {"x": 0, "y": 0})
            if not isinstance(self.latestPos, dict):
                self.latestPos = {"x": 0, "y": 0}
            self.root.geometry(f"800x300+{self.latestPos['x']}+{self.latestPos['y']}")
            self.theme_key = setting.get("Theme", self.theme_key)
            self.language_file = setting.get("Language", self.language_file)
            self.language_data = self.LoadLanguage(self.language_file)
            self.pending_load_file = setting.get("LoadFile")
            self.font_file = setting.get("Font", self.font_file)
            self.font = self.LoadFont(self.font_file)
            CONSOLE.print(f"[*] Settings loaded: Theme={self.theme_key}, Language={self.language_file}, Font='{self.font_file}'", style="#00bb00")
        except Exception as e:
            CONSOLE.print(f"[!] Failed to load settings: {e}", style="#bb0000")
            messagebox.showerror(self.tr("错误"), f"{self.tr('加载设置失败')}: {e}")

    @timer()
    def SaveSetting(self, key, value):
        try:
            if os.path.exists("assets/setting.json"):
                with open("assets/setting.json", "r", encoding="utf-8") as f:
                    setting = json.load(f)
                if not isinstance(setting, dict):
                    CONSOLE.print("[!] Invalid settings format while saving, creating new settings file.", style="#bbbb00")
                    setting = {"ScreenPos": self.latestPos, "LoadFile": None, "Theme": self.theme_key, "Language": self.language_file, "Font": self.font_file}
            else:
                setting = {"ScreenPos": self.latestPos, "LoadFile": None, "Theme": self.theme_key, "Language": self.language_file, "Font": self.font_file}
            setting[key] = value
            with open("assets/setting.json", "w", encoding="utf-8") as f:
                json.dump(setting, f, ensure_ascii=False, indent=2)

            CONSOLE.print(f"[*] Setting saved: {key} = {value}", style="#00bb00")
        except Exception as e:
            CONSOLE.print(f"[!] Failed to save setting: {e}", style="#bb0000")
            messagebox.showerror(self.tr("错误"), f"{self.tr('保存设置失败')}: {e}")

    def LoadLanguageFiles(self):
        languages = []
        try:
            for filename in os.listdir("assets/lang"):
                if filename.lower().endswith(".json"):
                    data = self.LoadLanguage(filename)
                    name = data.get("_name", os.path.splitext(filename)[0])
                    languages.append((filename, name))
        except Exception:
            pass
        return languages

    def LoadLanguage(self, filename):
        path = os.path.join("assets/lang", filename)
        try:
            with open(path, "r", encoding="utf-8-sig") as f:
                language_data = json.load(f)
                if not isinstance(language_data, dict):
                    return {}
                return language_data
        except Exception:
            return {}

    def ChangeLanguage(self, filename):
        if not filename:
            return
        self.language_file = filename
        self.language_data = self.LoadLanguage(filename)
        self.ApplyLanguage()
        self.SaveSetting("Language", filename)

    def ApplyLanguage(self):
        self.root.title(self.tr("choice UI"))
        if hasattr(self, 'showLabel'):
            self.showLabel.config(text=self.tr("等待抽取"))
        if hasattr(self, 'startButton'):
            self.startButton.config(text=self.tr("开始抽奖"))
        if hasattr(self, 'setButton'):
            self.setButton.config(text=self.tr("设置选项"))
        if hasattr(self, 'countInfoLabel'):
            self.countInfoLabel.config(text=self.tr("名单数量：{total}").format(total=len(self.choiseList)))
        self.UpdateLoadedFileLabel()
        if hasattr(self, 'MainMenu'):
            try:
                self.MainMenu.entryconfig(0, label=self.tr("软件信息"))
                self.MainMenu.entryconfig(1, label=self.tr("设置"))
                self.codeMenu.entryconfig(0, label=self.tr("源代码"))
                self.codeMenu.entryconfig(1, label=self.tr("关于"))
                self.SettingMenu.entryconfig(0, label=self.tr("主题"))
                self.SettingMenu.entryconfig(1, label=self.tr("语言"))
                self.SettingMenu.entryconfig(2, label=self.tr("字体"))
            except Exception as e:
                CONSOLE.print(f"[!] Failed to apply language: {e}", style="#bb0000")

    def ApplyFont(self):
        if hasattr(self, 'showLabel'):
            self.showLabel.config(font=(self.font, 20))
        if hasattr(self, 'countInfoLabel'):
            self.countInfoLabel.config(font=(self.font, 10))
        if hasattr(self, 'loadedFileLabel'):
            self.loadedFileLabel.config(font=(self.font, 10))
        if hasattr(self, 'fontLabel'):
            self.fontLabel.config(font=(self.font, 10))
        if hasattr(self, 'previewTitle'):
            self.previewTitle.config(font=(self.font, 10))
        if hasattr(self, 'previewLabel'):
            self.previewLabel.config(font=(self.font, 20))

    def GetFontDisplayName(self):
        if self.font_file:
            return os.path.basename(self.font_file)
        return self.font

    def UnregisterFontFile(self, font_path):
        if not font_path or not os.path.exists(font_path):
            return False
        if platform.system() != "Windows":
            return False
        try:
            FR_PRIVATE = 0x10
            result = ctypes.windll.gdi32.RemoveFontResourceExW(font_path, FR_PRIVATE, 0)
            return bool(result)
        except Exception:
            return False

    def LoadFont(self, font_value):
        if not font_value:
            CONSOLE.print("[-] No font specified, using default.", style="#bbbb00")
            return fontloader("Microsoft YaHei")
        if os.path.exists(font_value):
            CONSOLE.print(f"[*] Loaded font: {font_value}", style="#00bb00")
            return fontloader(font_value)
        return fontloader("Microsoft YaHei")

    def UpdateFontPreview(self, font_path=None):
        CONSOLE.print(f"[*] Loading font for preview: {font_path}", style="#00bb00")
        preview_family = self.LoadFont(font_path)
        CONSOLE.print(f"[*] Preview font family: {preview_family}", style="#00bb00")
        if hasattr(self, 'previewLabel'):
            self.previewLabel.config(text=self.tr("文本\nAaBbCc\n123"), font=(preview_family, 20))
            CONSOLE.print(f"[*] Updated font preview: {preview_family}", style="#00bb00")

    def OnFontSelectChange(self, event=None):
        selection = self.fontListbox.curselection()
        if not selection:
            return
        CONSOLE.print(f"[*] Font selection changed: index={selection[0]}", style="#00bb00")
        filename = self.fontListbox.get(selection[0])
        font_path = os.path.join(self.font_folder, filename)
        CONSOLE.print(f"[*] Selected font file: {font_path}", style="#00bb00")
        self.UpdateFontPreview(font_path)

    def ResolveTheme(self, theme_key=None):
        theme_key = theme_key or self.theme_key
        if theme_key == "system":
            system_theme = get_system_theme()
            if system_theme == "Dark":
                return self.themes.get("dark", self.themes["light"])
            return self.themes.get("light", self.themes["dark"])
        return self.themes.get(theme_key, self.themes["light"])

    def _is_dark_theme(self, theme):
        try:
            c = theme.get("bg", "#ffffff").lstrip('#')
            rr = int(c[0:2], 16); gg = int(c[2:4], 16); bb = int(c[4:6], 16)
            lum = (rr * 299 + gg * 587 + bb * 114) / 1000
            return lum < 128
        except Exception:
            return False

    def set_window_titlebar_color(self, theme):
        if platform.system() != "Windows":
            return
        try:
            hwnd = self.root.winfo_id()
            if not hwnd:
                return
            DWMWA_CAPTION_COLOR = 35
            DWMWA_TEXT_COLOR = 36
            DWMWA_USE_IMMERSIVE_DARK_MODE = 20
            dwmapi = ctypes.windll.dwmapi
            is_dark = self._is_dark_theme(theme)
            dark_value = ctypes.c_int(1 if is_dark else 0)
            try:
                dwmapi.DwmSetWindowAttribute(hwnd, DWMWA_USE_IMMERSIVE_DARK_MODE, ctypes.byref(dark_value), ctypes.sizeof(dark_value))
            except Exception:
                pass
            caption_color = int(theme.get("btn_bg", theme.get("bg", "#ffffff")).lstrip('#'), 16)
            text_color = int(theme.get("btn_fg", theme.get("fg", "#000000")).lstrip('#'), 16)
            caption_value = ctypes.c_int(caption_color)
            text_value = ctypes.c_int(text_color)
            try:
                dwmapi.DwmSetWindowAttribute(hwnd, DWMWA_CAPTION_COLOR, ctypes.byref(caption_value), ctypes.sizeof(caption_value))
            except Exception:
                pass
            try:
                dwmapi.DwmSetWindowAttribute(hwnd, DWMWA_TEXT_COLOR, ctypes.byref(text_value), ctypes.sizeof(text_value))
            except Exception:
                pass
        except Exception:
            pass

    def ApplyTheme(self):
        theme = self.ResolveTheme()
        self.style.configure("TLabel", background=theme["bg"], foreground=theme["fg"])
        self.style.configure("TButton", background=theme["btn_bg"], foreground=theme["btn_fg"], relief="flat", borderwidth=1)
        self.style.map(
            "TButton",
            background=[("!disabled", theme["btn_bg"]), ("active", theme["active_bg"])],
            foreground=[("!disabled", theme["btn_fg"])]
        )
        self.style.configure("TEntry", fieldbackground=theme["entry_bg"], foreground=theme["fg"], background=theme["entry_bg"])
        self.style.configure("TSpinbox", fieldbackground=theme["entry_bg"], foreground=theme["fg"], background=theme["entry_bg"])
        self.style.configure("TCombobox", fieldbackground=theme["entry_bg"], foreground=theme["fg"], background=theme["entry_bg"])
        self.style.configure("Green.Horizontal.TProgressbar", troughcolor=theme["entry_bg"], background="#22c55e", bordercolor=theme["border"], lightcolor="#4ade80", darkcolor="#166534", thickness=16)
        self.style.map("Green.Horizontal.TProgressbar", background=[("!disabled", "#22c55e")])
        self.style.configure("Treeview", background=theme["entry_bg"], fieldbackground=theme["entry_bg"], foreground=theme["fg"], bordercolor=theme["border"], rowheight=22)
        # Custom Treeview style for clearer grid-like appearance
        self.style.configure("Custom.Treeview", background=theme["entry_bg"], fieldbackground=theme["entry_bg"], foreground=theme["fg"], bordercolor=theme["border"], rowheight=22)
        # ensure selected state colors are mapped for the custom style
        try:
            self.style.map("Custom.Treeview", background=[("selected", theme.get("active_bg", theme["entry_bg"]))], foreground=[("selected", theme.get("btn_fg", theme["fg"]))])
        except Exception:
            pass
        self.style.configure("Treeview.Heading", background=theme.get("btn_bg", "#ffffff"), foreground=theme.get("btn_fg", theme["fg"]), relief="flat", font=(self.font, 10), borderwidth=1)
        self.style.map("Treeview.Heading", background=[("active", theme.get("active_bg", theme["entry_bg"]))])
        self.root.config(bg=theme["bg"])
        self.set_window_titlebar_color(theme)
        self.root.option_add("*Menu.background", theme["bg"])
        self.root.option_add("*Menu.foreground", theme["fg"])
        self.root.option_add("*Menu.activeBackground", theme["active_bg"])
        self.root.option_add("*Menu.activeForeground", theme["btn_fg"])
        self.root.option_add("*Menu.selectBackground", theme["active_bg"])
        self.root.option_add("*Menu.selectForeground", theme["btn_fg"])
        if hasattr(self, 'MainMenu'):
            try:
                self.MainMenu.config(bg=theme["bg"], fg=theme["fg"], activebackground=theme["active_bg"], activeforeground=theme["btn_fg"])
            except Exception:
                pass
        if hasattr(self, 'codeMenu'):
            try:
                self.codeMenu.config(bg=theme["bg"], fg=theme["fg"], activebackground=theme["active_bg"], activeforeground=theme["btn_fg"])
            except Exception:
                pass
        if hasattr(self, 'SettingMenu'):
            try:
                self.SettingMenu.config(bg=theme["bg"], fg=theme["fg"], activebackground=theme["active_bg"], activeforeground=theme["btn_fg"])
            except Exception:
                pass
        self.apply_theme_to_window(self.root)

    def apply_theme_to_window(self, window):
        theme = self.themes.get(self.theme_key, self.themes["light"])
        for child in window.winfo_children():
            try:
                if isinstance(child, tk.Label):
                    child.config(bg=theme["bg"], fg=theme["fg"])
                elif isinstance(child, tk.Frame):
                    child.config(bg=theme["bg"])
                elif isinstance(child, tk.Listbox):
                    child.config(bg=theme["entry_bg"], fg=theme["fg"], highlightbackground=theme["border"], selectbackground=theme["active_bg"], selectforeground=theme["btn_fg"])
                elif isinstance(child, tk.Entry) or isinstance(child, tk.Text):
                    child.config(bg=theme["entry_bg"], fg=theme["fg"], insertbackground=theme["fg"], highlightbackground=theme["border"])
                elif isinstance(child, tk.Spinbox):
                    child.config(bg=theme["entry_bg"], fg=theme["fg"], highlightbackground=theme["border"])
                elif isinstance(child, tk.Menu):
                    child.config(bg=theme["bg"], fg=theme["fg"], activebackground=theme["active_bg"], activeforeground=theme["btn_fg"])
                elif isinstance(child, ttk.Combobox):
                    child.configure(style="TCombobox")
                elif isinstance(child, ttk.Button):
                    child.configure(style="TButton")
                elif isinstance(child, ttk.Entry):
                    child.configure(style="TEntry")
                elif isinstance(child, ttk.Spinbox):
                    child.configure(style="TSpinbox")
                elif isinstance(child, ttk.Treeview):
                    child.configure(style="Custom.Treeview")
                    try:
                        child.configure(background=theme["entry_bg"], fieldbackground=theme["entry_bg"], foreground=theme["fg"])
                    except Exception:
                        pass
                if isinstance(child, tk.Toplevel):
                    child.config(bg=theme["bg"])
                self.apply_theme_to_window(child)
            except Exception:
                pass

    def tr(self, text):
        CONSOLE.print(f"[*] Translating text: '{text}'", style="#00bb00")
        if not isinstance(text, str):
            return text
        if self.language_data and text in self.language_data:
            return self.language_data[text]
        return text

    def LoadChoiceFiles(self):
        files = []
        try:
            for name in os.listdir(self.choice_folder):
                if name.lower().endswith(".json"):
                    files.append(name)
        except Exception:
            pass
        return files

    def UpdateCountLimit(self):
        total = len(self.choiseList)
        maxval = max(1, min(100, total))
        self.countSpinBox.config(to=maxval)
        if self.countVar.get() > maxval:
            self.countVar.set(maxval)
        self.countInfoLabel.config(text=self.tr("名单数量：{total}").format(total=total))
        self.UpdateLoadedFileLabel()
        if total == 0:
            self.startButton.state(["disabled"])
        else:
            self.startButton.state(["!disabled"])

    def UpdateLoadedFileLabel(self):
        if self.current_file_path:
            self.loadedFileLabel.config(text=self.tr("当前名单：{name}").format(name=os.path.basename(self.current_file_path)))
        else:
            self.loadedFileLabel.config(text=self.tr("当前名单：无"))

    @timer()
    def LoadFile(self, file_path=None):
        if not file_path:
            print('None File.')
            return
        try:
            if file_path.lower().endswith(".txt"):
                self.ImportTxtFile(file_path)
                return
            if file_path.lower().endswith(".json"):
                with open(file_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                if not isinstance(data, list):
                    raise ValueError(self.tr("名单文件格式不正确，应为名称列表。"))
                self.choiseList = self.NormalizeNames(data)
                if not self.choiseList:
                    messagebox.showwarning(self.tr("提示"), self.tr("名单文件为空。"))
                self.current_file_path = file_path
                self.UpdateCountLimit()
                self.SaveSetting("LoadFile", file_path)
                return
            messagebox.showwarning(self.tr("提示"), self.tr("仅支持 TXT 或 JSON 文件导入。"))
        except Exception as e:
            CONSOLE.print(f"[!] Failed to load file: {e}", style="#bb0000")
            messagebox.showerror(self.tr("错误"), f"{self.tr('加载文件失败')}: {e}")
            return

    @timer()
    def ImportTxtFile(self, file_path):
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                lines = [line.strip() for line in f if line.strip()]
            if not lines:
                messagebox.showwarning(self.tr("提示"), self.tr("文本文件没有可用名单内容。"))
                return
            names = self.NormalizeNames(lines)
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            target_name = base_name + ".json"
            target_path = os.path.join(self.choice_folder, target_name)
            count = 1
            while os.path.exists(target_path):
                target_name = f"{base_name}_{count}.json"
                target_path = os.path.join(self.choice_folder, target_name)
                count += 1
            with open(target_path, "w", encoding="utf-8") as f:
                json.dump(names, f, ensure_ascii=False, indent=2)
            self.choiseList = names
            self.current_file_path = target_path
            self.UpdateCountLimit()
            self.SaveSetting("LoadFile", target_path)
            message = self.tr("已导入 {count} 条名单，并保存为：{path}").format(count=len(names), path=target_path)
            if len(names) < len(lines):
                message += "\n" + self.tr("已自动去重，去除重复项 {removed} 条。").format(removed=len(lines) - len(names))
            messagebox.showinfo(self.tr("导入成功"), message)
        except Exception as e:
            CONSOLE.print(f"[!] Failed to import TXT: {e}", style="#bb0000")
            messagebox.showerror(self.tr("错误"), f"{self.tr('导入 TXT 失败')}: {e}")
            return

    @timer()
    def LoadNamesFromJson(self, file_path):
        if not file_path:
            raise ValueError("无效的名单文件路径。")
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, list):
            raise ValueError("名单文件格式不正确，应为名称列表。")
        return self.NormalizeNames(data)

    def NormalizeNames(self, names):
        seen = set()
        normalized = []
        for item in names:
            name = str(item).strip()
            if not name or name in seen:
                continue
            seen.add(name)
            normalized.append(name)
        return normalized

    def SaveNamesToJson(self, file_path, names):
        if not file_path:
            raise ValueError("目标文件路径无效，无法保存名单。")
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(names, f, ensure_ascii=False, indent=2)

    def OpenFileEditor(self, file_path):
        FileEditor(self, file_path)

    def UIposChanged(self):
        while True:
            if self.WindowMove():
                x = self.root.winfo_x()
                y = self.root.winfo_y()
                if self.latestPos!= {"x": x, "y": y}:
                    self.SaveSetting("ScreenPos", {"x": x, "y": y})
                    self.latestPos = {"x": x, "y": y}
                CONSOLE.print("[*] Window position changed: ", self.latestPos, style="#00bb00")
            time.sleep(0.1)
    
    def WindowMove(self):
        x = self.root.winfo_x()
        y = self.root.winfo_y()
        if {"x": x, "y": y} != self.latestPos:
            return True
        return False

class FileEditor:
    def __init__(self, parent, file_path):
        self.parent = parent
        if not file_path:
            messagebox.showerror(self.parent.tr("错误"), self.parent.tr("未提供名单文件路径，无法打开编辑器。"))
            return
        self.file_path = os.path.abspath(file_path)
        self.names = self.parent.LoadNamesFromJson(self.file_path)

        self.editor = tk.Toplevel(parent.root)
        self.editor.title(self.parent.tr("编辑名单 - {name}").format(name=os.path.basename(file_path)))
        self.editor.resizable(False, False)
        self.editor.iconbitmap("assets/icon.ico")
        self.editor.geometry("600x450")
        self.editor.config(bg=self.parent.themes[self.parent.theme_key]["bg"])
        self.SetControls()
        self.RefreshNames()
        self.parent.apply_theme_to_window(self.editor)
        self.editor.grab_set()

    def SetControls(self):
        title_label = tk.Label(self.editor, text=self.parent.tr("名单内容"), font=(self.parent.font, 12))
        title_label.place(x=20, y=10)

        self.nameListbox = tk.Listbox(self.editor, height=14, width=40)
        self.nameListbox.place(x=20, y=40)

        self.addEntry = ttk.Entry(self.editor, width=30)
        self.addEntry.place(x=20, y=340)

        self.addButton = ttk.Button(self.editor, text=self.parent.tr("添加"), command=self.AddName)
        self.addButton.place(x=300, y=340, width=80)

        self.deleteButton = ttk.Button(self.editor, text=self.parent.tr("删除选中"), command=self.DeleteSelected)
        self.deleteButton.place(x=20, y=400, width=120)

        self.saveButton = ttk.Button(self.editor, text=self.parent.tr("保存修改"), command=self.SaveNames)
        self.saveButton.place(x=150, y=400, width=120)

        self.closeButton = ttk.Button(self.editor, text=self.parent.tr("关闭"), command=self.editor.destroy)
        self.closeButton.place(x=280, y=400, width=120)

        self.addrange = ttk.Button(self.editor, text=self.parent.tr("添加序列"), command=self.ImportRange)
        self.addrange.place(x=410, y=400, width=120)

        self.importExcelButton = ttk.Button(self.editor, text=self.parent.tr("导入Excel"), command=self.ImportExcel)
        self.importExcelButton.place(x=450, y=240, width=100)

    def ImportRange(self):
        startnum = simpledialog.askinteger(self.parent.tr("添加序列"), self.parent.tr("请输入起始编号："), parent=self.editor)
        endnum = simpledialog.askinteger(self.parent.tr("添加序列"), self.parent.tr("请输入结束编号："), parent=self.editor)
        if startnum is None or endnum is None:
            return
        if startnum > endnum:
            messagebox.showwarning(self.parent.tr("提示"), self.parent.tr("起始编号必须小于等于结束编号。"))
            return
        for i in range(startnum, endnum+1):
            name = str(i).strip()
            if not name:
                continue
            if name in self.names:
                continue
            self.names.append(name)

        self.RefreshNames()

    def ImportExcel(self):
        path = filedialog.askopenfilename(
            title=self.parent.tr("选择 Excel 文件"),
            filetypes=[("Excel 文件", "*.xlsx *.xls *.xlsm *.xltx *.xltm"), ("所有文件", "*")]
        )
        if not path:
            return
        try:
            workbook = self._open_excel_workbook(path)
            sheet_names = []
            if hasattr(workbook, "sheetnames"):
                sheet_names = list(workbook.sheetnames)
            elif hasattr(workbook, "sheet_names"):
                sheet_names = list(workbook.sheet_names())
            if not sheet_names:
                raise ValueError(self.parent.tr("Excel 文件中未找到工作表。"))
            sheet_name = self._choose_excel_sheet(sheet_names)
            if not sheet_name:
                return
            selected_values = self._show_excel_sheet_selector(workbook, sheet_name, path)
            if hasattr(workbook, "close"):
                try:
                    workbook.close()
                except Exception:
                    pass
            if not selected_values:
                messagebox.showwarning(self.parent.tr("提示"), self.parent.tr("未选择任何单元格。"))
                return
            imported_names = self.parent.NormalizeNames(selected_values)
            if not imported_names:
                messagebox.showwarning(self.parent.tr("提示"), self.parent.tr("所选单元格中没有有效姓名数据。"))
                return
            added_count = 0
            for name in imported_names:
                if name not in self.names:
                    self.names.append(name)
                    added_count += 1
            self.RefreshNames()
            messagebox.showinfo(
                self.parent.tr("导入成功"),
                self.parent.tr("已导入 {count} 条数据，新增 {added} 条。").format(count=len(imported_names), added=added_count)
            )
        except Exception as e:
            CONSOLE.print(f"[!] Failed to import Excel: {e}", style="#bb0000")
            messagebox.showerror(self.parent.tr("错误"), f"{self.parent.tr('导入 Excel 失败')}: {e}")

    def _show_excel_sheet_selector(self, workbook, sheet_name, path):
        sheet_values = self._extract_sheet_values(workbook, sheet_name, path)
        if not sheet_values:
            raise ValueError(self.parent.tr("工作表中未找到任何数据。"))

        max_rows = min(len(sheet_values), 100)
        max_cols = min(max((len(row) for row in sheet_values), default=0), 20)
        display_rows = sheet_values[:max_rows]

        selector = tk.Toplevel(self.editor)
        selector.title(self.parent.tr("选择单元格 - {sheet}").format(sheet=sheet_name))
        selector.resizable(False, False)
        selector.geometry("900x900")
        selector.transient(self.editor)
        selector.iconbitmap("assets/icon.ico")
        selector.grab_set()
        try:
            selector.config(bg=self.parent.themes[self.parent.theme_key]["bg"])
        except Exception:
            pass
        try:
            self.parent.apply_theme_to_window(selector)
        except Exception:
            pass

        theme = self.parent.themes[self.parent.theme_key]
        selector_style = ttk.Style(selector)
        try:
            selector_style.configure("Selector.TButton", background=theme["btn_bg"], foreground=theme["btn_fg"], relief="flat", borderwidth=1)
            selector_style.map("Selector.TButton", background=[("active", theme["active_bg"])], foreground=[("active", theme["btn_fg"])])
            selector_style.configure("Selector.TEntry", fieldbackground=theme["entry_bg"], background=theme["entry_bg"], foreground=theme["fg"], bordercolor=theme["border"])
            selector_style.map("Selector.TEntry", fieldbackground=[("disabled", theme["entry_bg"])], foreground=[("disabled", theme["fg"])])
            selector_style.configure("Selector.Treeview", background=theme["entry_bg"], fieldbackground=theme["entry_bg"], foreground=theme["fg"])
        except Exception:
            pass

        tk.Label(selector, text=self.parent.tr("工作表：{sheet}，已显示前 {rows} 行、{cols} 列。")
                 .format(sheet=sheet_name, rows=max_rows, cols=max_cols), font=(self.parent.font, 11), bg=theme["bg"], fg=theme["fg"]).pack(pady=8)

        table_frame = tk.Frame(selector, bg=theme["bg"])
        table_frame.pack(fill="both", expand=True, padx=8)

        cols = [self._col_letter(i + 1) for i in range(max_cols)]
        tree = ttk.Treeview(table_frame, columns=cols, show="headings", height=18, style="Selector.Treeview")
        for col in cols:
            tree.heading(col, text=col)
            tree.column(col, width=110, anchor="w")

        try:
            theme = self.parent.themes[self.parent.theme_key]
            entry_bg = theme.get("entry_bg", "#ffffff")
            border = theme.get("border", "#f7f7f7")
            def _shade_color(hex_color, factor):
                try:
                    c = hex_color.lstrip('#')
                    r = int(c[0:2], 16)
                    g = int(c[2:4], 16)
                    b = int(c[4:6], 16)
                    r = max(0, min(255, int(r * factor)))
                    g = max(0, min(255, int(g * factor)))
                    b = max(0, min(255, int(b * factor)))
                    return f"#{r:02x}{g:02x}{b:02x}"
                except Exception:
                    return hex_color

            c = entry_bg.lstrip('#')
            try:
                rr = int(c[0:2], 16); gg = int(c[2:4], 16); bb = int(c[4:6], 16)
                lum = (rr * 299 + gg * 587 + bb * 114) / 1000
            except Exception:
                lum = 200
            factor = 0.96 if lum > 128 else 1.04
            odd_bg = entry_bg
            even_bg = _shade_color(entry_bg, factor)
            tree.tag_configure('odd', background=odd_bg, foreground=theme.get('fg', '#000000'))
            tree.tag_configure('even', background=even_bg, foreground=theme.get('fg', '#000000'))
            tree.tag_configure('selected', background=theme.get('active_bg', '#e5e7eb'), foreground=theme.get('btn_fg', theme.get('fg', '#000000')))
            tree.configure(background=entry_bg, fieldbackground=entry_bg, foreground=theme.get('fg', '#000000'))
        except Exception:
            pass

        for row_index, row in enumerate(display_rows, start=1):
            values = [self._format_excel_value(row[col_index]) if col_index < len(row) else "" for col_index in range(max_cols)]
            tag = 'even' if row_index % 2 == 0 else 'odd'
            tree.insert("", "end", iid=str(row_index), values=values, tags=(tag,))

        try:
            font_obj = tkfont.Font(font=(self.parent.font, 10))
            padding = 18
            for ci, col in enumerate(cols):
                max_w = font_obj.measure(col)
                for row in display_rows:
                    if ci < len(row):
                        text = self._format_excel_value(row[ci])
                        w = font_obj.measure(text)
                        if w > max_w:
                            max_w = w
                tree.column(col, width=max_w + padding)
        except Exception:
            pass

        def _apply_selection_tags(event=None):
            try:
                sel = set(tree.selection())
                for iid in tree.get_children(''):
                    tags = [t for t in tree.item(iid, 'tags') if t != 'selected']
                    if iid in sel:
                        tags.append('selected')
                    tree.item(iid, tags=tuple(tags))
            except Exception:
                pass

        tree.bind('<<TreeviewSelect>>', _apply_selection_tags)
        _apply_selection_tags()

        vscroll = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        hscroll = ttk.Scrollbar(table_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vscroll.set, xscrollcommand=hscroll.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vscroll.grid(row=0, column=1, sticky="ns")
        hscroll.grid(row=1, column=0, sticky="ew")
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        range_frame = tk.Frame(selector, bg=theme["bg"])
        range_frame.pack(fill="x", padx=8, pady=4)
        tk.Label(range_frame, text=self.parent.tr("范围地址："), font=(self.parent.font, 11), bg=theme["bg"], fg=theme["fg"]).grid(row=0, column=0, sticky="w")
        start_entry = ttk.Entry(range_frame, width=12, style="Selector.TEntry")
        start_entry.grid(row=0, column=1, padx=4)
        tk.Label(range_frame, text="-", font=(self.parent.font, 11), bg=theme["bg"], fg=theme["fg"]).grid(row=0, column=2)
        end_entry = ttk.Entry(range_frame, width=12, style="Selector.TEntry")
        end_entry.grid(row=0, column=3, padx=4)
        ttk.Button(range_frame, text=self.parent.tr("添加范围"), command=lambda: add_range_selection(), style="Selector.TButton").grid(row=0, column=4, padx=12)
        tk.Label(range_frame, text=self.parent.tr("示例: A1 D5 或 A1:D5"), font=(self.parent.font, 9), bg=theme["bg"], fg=theme["fg"]).grid(row=1, column=0, columnspan=5, sticky="w", pady=4)
        shift_status_label = tk.Label(range_frame, text=self.parent.tr("按住 Shift 双击两个角点选择区域。"), font=(self.parent.font, 9), fg=theme["fg"], bg=theme["bg"])
        shift_status_label.grid(row=2, column=0, columnspan=5, sticky="w", pady=2)

        selection_frame = tk.Frame(selector, bg=theme["bg"])
        selection_frame.pack(fill="x", padx=8, pady=8)

        tk.Label(selection_frame, text=self.parent.tr("已选单元格："), font=(self.parent.font, 11), bg=theme["bg"], fg=theme["fg"]).pack(anchor="w")
        selected_listbox = tk.Listbox(selection_frame, height=5, bg=theme["entry_bg"], fg=theme["fg"], highlightbackground=theme["border"], selectbackground=theme["active_bg"], selectforeground=theme["btn_fg"])
        selected_listbox.pack(fill="x", padx=4, pady=4)

        selected_values = []
        selected_set = set()
        shift_anchor = None

        try:
            start_entry.config(background=theme["entry_bg"], foreground=theme["fg"], insertbackground=theme["fg"])
            end_entry.config(background=theme["entry_bg"], foreground=theme["fg"], insertbackground=theme["fg"])
            selected_listbox.config(bg=theme["entry_bg"], fg=theme["fg"], highlightbackground=theme["border"], selectbackground=theme["active_bg"], selectforeground=theme["btn_fg"])
        except Exception:
            pass

        def add_cell_selection(event=None):
            nonlocal shift_anchor
            row_id = tree.identify_row(event.y) if event else None
            col_id = tree.identify_column(event.x) if event else None
            if not row_id or not col_id:
                return
            col_index = int(col_id.replace("#", "")) - 1
            row_index = int(row_id) - 1
            if row_index >= len(display_rows) or col_index >= len(display_rows[row_index]):
                return
            address = f"{self._col_letter(col_index + 1)}{row_index + 1}"
            is_shift = event and (event.state & 0x0001 != 0)
            if is_shift:
                if shift_anchor is None:
                    shift_anchor = (row_index, col_index)
                    shift_status_label.config(text=f"Shift 锚点已设为 {address}，再按住 Shift 双击第二个角点。")
                    return
                add_range_by_coords(shift_anchor[0], shift_anchor[1], row_index, col_index)
                shift_anchor = None
                shift_status_label.config(text=self.parent.tr("按住 Shift 双击两个角点选择区域。"))
                return
            value = display_rows[row_index][col_index]
            text = self._format_excel_value(value)
            if not text:
                return
            add_selection_item(address, text)

        def add_selection_item(address, text):
            item = f"{address}: {text}"
            if item in selected_set:
                return
            selected_set.add(item)
            selected_values.append(text)
            selected_listbox.insert(tk.END, item)

        def add_range_by_coords(r1, c1, r2, c2):
            if r1 > r2:
                r1, r2 = r2, r1
            if c1 > c2:
                c1, c2 = c2, c1
            for row_index in range(r1, r2 + 1):
                if row_index < 0 or row_index >= len(display_rows):
                    continue
                for col_index in range(c1, c2 + 1):
                    if col_index < 0 or col_index >= max_cols:
                        continue
                    value = display_rows[row_index][col_index]
                    text = self._format_excel_value(value)
                    if not text:
                        continue
                    address = f"{self._col_letter(col_index + 1)}{row_index + 1}"
                    add_selection_item(address, text)

        def parse_cell_address(address_text):
            address_text = address_text.strip().upper().replace(" ", "")
            if not address_text:
                return None
            if ":" in address_text:
                return None
            letters = ''.join(ch for ch in address_text if ch.isalpha())
            digits = ''.join(ch for ch in address_text if ch.isdigit())
            if not letters or not digits:
                return None
            col = 0
            for ch in letters:
                col = col * 26 + (ord(ch) - 64)
            row = int(digits)
            return (row - 1, col - 1)

        def parse_range_text(text_start, text_end):
            text_start = text_start.strip().upper().replace(" ", "")
            text_end = text_end.strip().upper().replace(" ", "")
            if not text_start or not text_end:
                return None
            if ":" in text_start and ":" not in text_end:
                parts = text_start.split(":")
                if len(parts) == 2:
                    text_start, text_end = parts[0], parts[1]
            start = parse_cell_address(text_start)
            end = parse_cell_address(text_end)
            if not start or not end:
                return None
            r1, c1 = start
            r2, c2 = end
            return (min(r1, r2), min(c1, c2), max(r1, r2), max(c1, c2))

        def add_range_selection():
            range_text = start_entry.get().strip()
            end_text = end_entry.get().strip()
            range_coords = parse_range_text(range_text, end_text)
            if not range_coords:
                messagebox.showwarning(self.parent.tr("提示"), self.parent.tr("请输入有效的单元格范围，例如 A1:D5 或 A1 D5。"), parent=selector)
                return
            r1, c1, r2, c2 = range_coords
            for row_index in range(r1, r2 + 1):
                if row_index < 0 or row_index >= len(display_rows):
                    continue
                for col_index in range(c1, c2 + 1):
                    if col_index < 0 or col_index >= max_cols:
                        continue
                    value = display_rows[row_index][col_index]
                    text = self._format_excel_value(value)
                    if not text:
                        continue
                    address = f"{self._col_letter(col_index + 1)}{row_index + 1}"
                    add_selection_item(address, text)

        def remove_selection():
            selection = selected_listbox.curselection()
            if not selection:
                return
            for index in reversed(selection):
                item = selected_listbox.get(index)
                selected_listbox.delete(index)
                selected_set.discard(item)
                value = item.split(": ", 1)[1]
                selected_values.remove(value)

        tree.bind("<Double-1>", add_cell_selection)

        buttons_frame = tk.Frame(selector, bg=theme["bg"])
        buttons_frame.pack(pady=6)
        ttk.Button(buttons_frame, text=self.parent.tr("移除选中"), command=remove_selection, style="Selector.TButton").pack(side="left", padx=6)

        result = {"values": None}

        def confirm():
            result["values"] = selected_values.copy()
            selector.destroy()

        def cancel():
            selector.destroy()

        ttk.Button(buttons_frame, text=self.parent.tr("导入所选单元格"), command=confirm, style="Selector.TButton").pack(side="left", padx=6)
        ttk.Button(buttons_frame, text=self.parent.tr("取消"), command=cancel, style="Selector.TButton").pack(side="left", padx=6)

        self.editor.wait_window(selector)
        return result["values"]

    def _extract_sheet_values(self, workbook, sheet_name, path):
        values = []
        if path.lower().endswith(".xlsx") or path.lower().endswith(".xlsm") or path.lower().endswith(".xltx") or path.lower().endswith(".xltm"):
            worksheet = workbook[sheet_name]
            for row in worksheet.iter_rows(values_only=True):
                values.append(list(row))
        else:
            sheet = workbook.sheet_by_name(sheet_name)
            for row_index in range(sheet.nrows):
                values.append(list(sheet.row_values(row_index)))
        return values

    def _format_excel_value(self, value):
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _col_letter(self, col_index):
        letters = ""
        while col_index > 0:
            col_index, remainder = divmod(col_index - 1, 26)
            letters = chr(65 + remainder) + letters
        return letters

    def _open_excel_workbook(self, path):
        ext = os.path.splitext(path)[1].lower()
        if ext == ".xlsx" or ext == ".xlsm" or ext == ".xltx" or ext == ".xltm":
            return load_workbook(path, read_only=True, data_only=True)
        if ext == ".xls":
            return xlrd.open_workbook(path, on_demand=True)
        raise ValueError(self.parent.tr("仅支持 XLS/XLSX/XLSM/XLTX/XLTM 格式的 Excel 文件。"))

    def _choose_excel_sheet(self, sheet_names):
        if not sheet_names:
            return None
        if len(sheet_names) == 1:
            return sheet_names[0]

        picker = tk.Toplevel(self.editor)
        picker.title(self.parent.tr("选择工作表"))
        picker.resizable(False, False)
        picker.geometry("350x300")
        picker.transient(self.editor)
        picker.iconbitmap("assets/icon.ico")
        picker.grab_set()
        
        theme = self.parent.themes[self.parent.theme_key]
        try:
            picker.config(bg=theme["bg"])
        except Exception as e:
            CONSOLE.print(f"[!] Failed to set picker background: {e}", style="#bb0000")
        try:
            self.parent.apply_theme_to_window(picker)
        except Exception as e:
            CONSOLE.print(f"[!] Failed to apply theme to picker: {e}", style="#bb0000")

        picker_style = ttk.Style(picker)
        try:
            picker_style.configure("Picker.TButton", background=theme["btn_bg"], foreground=theme["btn_fg"], relief="flat", borderwidth=1)
            picker_style.map("Picker.TButton", background=[("active", theme["active_bg"])], foreground=[("active", theme["btn_fg"])])
        except Exception:
            pass

        label = tk.Label(picker, text=self.parent.tr("请选择要导入的工作表："), font=(self.parent.font, 11), bg=theme["bg"], fg=theme["fg"])
        label.pack(pady=12)

        listbox = tk.Listbox(picker, height=8, width=38, bg=theme["entry_bg"], fg=theme["fg"], highlightbackground=theme["border"], selectbackground=theme["active_bg"], selectforeground=theme["btn_fg"])
        listbox.pack(padx=12)
        for name in sheet_names:
            listbox.insert(tk.END, name)

        result = {"sheet": None}

        def choose():
            selection = listbox.curselection()
            if not selection:
                messagebox.showwarning(self.parent.tr("提示"), self.parent.tr("请先选择一个工作表。"), parent=picker)
                return
            result["sheet"] = listbox.get(selection[0])
            picker.destroy()

        def cancel():
            picker.destroy()

        button_frame = tk.Frame(picker, bg=theme["bg"])
        button_frame.pack(pady=12)
        ttk.Button(button_frame, text=self.parent.tr("确定"), command=choose, style="Picker.TButton").pack(side="left", padx=8)
        ttk.Button(button_frame, text=self.parent.tr("取消"), command=cancel, style="Picker.TButton").pack(side="left", padx=8)
        listbox.bind("<Double-Button-1>", lambda event: choose())

        self.editor.wait_window(picker)
        return result["sheet"]

    def _read_excel_sheet(self, workbook, sheet_name, path):
        names = []
        if path.lower().endswith(".xlsx") or path.lower().endswith(".xlsm") or path.lower().endswith(".xltx") or path.lower().endswith(".xltm"):
            worksheet = workbook[sheet_name]
            for row in worksheet.iter_rows(values_only=True):
                value = self._first_nonempty_cell(row)
                if value is not None:
                    names.append(str(value).strip())
        else:
            sheet = workbook.sheet_by_name(sheet_name)
            for row_index in range(sheet.nrows):
                row = sheet.row_values(row_index)
                value = self._first_nonempty_cell(row)
                if value is not None:
                    names.append(str(value).strip())
        return names

    def _first_nonempty_cell(self, row):
        for cell in row:
            if cell is None:
                continue
            text = str(cell).strip()
            if text:
                return text
        return None

    def RefreshNames(self):
        self.nameListbox.delete(0, tk.END)
        for name in self.names:
            self.nameListbox.insert(tk.END, name)
        CONSOLE.print(f"[*] Names refreshed: {len(self.names)}", style="#00bb00")

    def AddName(self):
        new_name = self.addEntry.get().strip()
        if not new_name:
            CONSOLE.print("[-] No name provided for addition.", style="#bbbb00")
            messagebox.showwarning(self.parent.tr("提示"), self.parent.tr("请输入要添加的姓名。"))
            return
        self.names.append(new_name)
        self.addEntry.delete(0, tk.END)
        self.RefreshNames()

    def DeleteSelected(self):
        selection = self.nameListbox.curselection()
        if not selection:
            CONSOLE.print("[-] No name selected for deletion.", style="#bbbb00")
            messagebox.showwarning(self.parent.tr("提示"), self.parent.tr("请先选择一个要删除的姓名。"))
            return
        for index in reversed(selection):
            self.names.pop(index)
        self.RefreshNames()

    def SaveNames(self):
        if not self.file_path:
            CONSOLE.print("[-] No file path provided, cannot save names.", style="#bbbb00")
            messagebox.showerror(self.parent.tr("错误"), self.parent.tr("当前名单文件路径无效，请重新打开名单文件后再保存。"))
            return
        try:
            self.parent.SaveNamesToJson(self.file_path, self.names)
            if hasattr(self.parent, 'current_file_path') and os.path.abspath(self.file_path) == os.path.abspath(self.parent.current_file_path):
                self.parent.choiseList = self.names
                self.parent.UpdateCountLimit()
            CONSOLE.print(f"[*] Names saved: {len(self.names)}", style="#00bb00")
            messagebox.showinfo(self.parent.tr("保存成功"), self.parent.tr("名单已保存。"))
        except Exception as e:
            CONSOLE.print(f"[!] Failed to save names: {e}", style="#bb0000")
            messagebox.showerror(self.parent.tr("错误"), f"{self.parent.tr('保存名单失败')}: {e}")

if __name__ == "__main__":
    app = MainUI()